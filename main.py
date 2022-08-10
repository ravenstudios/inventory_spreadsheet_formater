import re
import pyexcel
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import requests


def get_data(data):
    pattern = "<td>\d{1,3}<\/td><td>(iPhone)(\s)?(X?|XR?|XS?)(\d{1,2})?(S)?(\+?)\s?(mini)?(Pro)?\s?(Max)?<\/td><td>LCD<\/td><td>(Black?|White)<\/td><td>AFTERMARKET<\/td><td>(\d{1,3})"
    filtered_list = []

    web_data = re.findall(pattern, data)
    for line in list(web_data):
        data = []
        name = ""

        for i in range(7):
            name += line[i]

        data.append(name)
        data.append(line[9])
        data.append(line[10])

        filtered_list.append(data)
    return filtered_list



def get_inventory_from_site():

    URL = "http://www.254repair.com/repairs//inshop2/inventoryreportapple.php"
    page = requests.get(URL)






# def get_xmls():
    workbook = Workbook()
    ws = workbook.active

    data = get_data(page.text)
    ws.column_dimensions['A'].width = 15
    ws.append(["Phone", "Color", "Quanity", "Front", "Back", "Diff"])
    for d in data:
        ws.append(d)

    row_count = ws.max_row
    yellow = "00FFFF00"
    for rows in ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=6):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=yellow, end_color=yellow, fill_type = "solid")
    workbook.save("inventory.xlsx")





get_inventory_from_site()
